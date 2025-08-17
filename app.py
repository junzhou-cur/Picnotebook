import os
import uuid
from datetime import datetime, timezone
# Test comment to demonstrate diff view functionality
from flask import Flask, request, render_template, jsonify, redirect, url_for, flash, session, current_app, send_file
from flask_login import LoginManager, login_required, current_user
from flask_jwt_extended import JWTManager, jwt_required, get_jwt_identity
from flask_migrate import Migrate
from flask_cors import CORS
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
from celery_app import make_celery
from config import config
from models import db, User, Project, LabNote, ProcessingJob, ProjectMember, Protocol, Sequence, Amplicon
from auth import auth_bp
from tasks import process_image_async, generate_project_report_async
from enhanced_ocr import EnhancedOCR, extract_text_from_image, process_lab_note_image  # Enhanced OCR integration
from lab_note_parser import LabNoteParser, process_lab_note_text  # Lab note parsing
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import io

def create_app(config_name=None):
    """Application factory"""
    app = Flask(__name__)
    
    # Load configuration
    config_name = config_name or os.environ.get('FLASK_ENV', 'development')
    app.config.from_object(config[config_name])
    
    # Initialize extensions
    db.init_app(app)
    
    # Initialize Flask-Mail
    mail = Mail(app)
    
    # Helper function to send email notifications
    def send_email(to, subject, template, **kwargs):
        try:
            msg = Message(
                subject=subject,
                recipients=[to] if isinstance(to, str) else to,
                sender=app.config['MAIL_DEFAULT_SENDER']
            )
            msg.html = template.format(**kwargs)
            mail.send(msg)
            return True
        except Exception as e:
            current_app.logger.error(f"Failed to send email: {str(e)}")
            return False
    
    # Manual CORS setup - bypassing Flask-CORS for development
    def add_cors_headers(response):
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
        return response
    
    app.after_request(add_cors_headers)
    
    # Handle preflight OPTIONS requests
    @app.before_request
    def handle_preflight():
        if request.method == "OPTIONS":
            response = jsonify({})
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
            return response
    
    # Flask-Login setup
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'auth.login'
    login_manager.login_message = 'Please log in to access this page.'
    login_manager.login_message_category = 'info'
    
    @login_manager.user_loader
    def load_user(user_id):
        return User.query.get(int(user_id))
    
    # JWT setup
    jwt = JWTManager(app)
    
    # Database migrations
    migrate = Migrate(app, db)
    
    # Celery setup
    celery = make_celery(app)
    app.celery = celery
    
    # Register blueprints
    app.register_blueprint(auth_bp, url_prefix='/auth')
    
    # Ensure upload directories exist
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    os.makedirs(app.config['PROJECTS_FOLDER'], exist_ok=True)
    
    # Direct image processing function
    def process_image_direct(image_path, api_key, user_id, project_id=None):
        """Process image directly without Celery"""
        try:
            import base64
            import requests
            
            # Read and encode image
            with open(image_path, 'rb') as img_file:
                image_data = base64.b64encode(img_file.read()).decode('utf-8')
            
            # Call xAI API
            headers = {
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json'
            }
            
            # Try different model names and message formats
            try:
                # First try with vision model
                payload = {
                    'model': 'grok-2-vision-1212',  # Updated model name
                    'messages': [
                        {
                            'role': 'user',
                            'content': [
                                {
                                    'type': 'text',
                                    'text': 'Please transcribe and organize this handwritten lab note into a structured markdown format. Include the date, experiment details, observations, and any conclusions. Extract all text accurately and organize it clearly.'
                                },
                                {
                                    'type': 'image_url',
                                    'image_url': {
                                        'url': f'data:image/jpeg;base64,{image_data}'
                                    }
                                }
                            ]
                        }
                    ],
                    'max_tokens': 2000,
                    'temperature': 0.3
                }
            except:
                # Fallback to text-only model
                payload = {
                    'model': 'grok-beta',
                    'messages': [
                        {
                            'role': 'user',
                            'content': 'I have uploaded a handwritten lab note image. Since you cannot see images, please provide a template for transcribing lab notes in structured markdown format.'
                        }
                    ],
                    'max_tokens': 1000,
                    'temperature': 0.3
                }
            
            response = requests.post(
                'https://api.x.ai/v1/chat/completions',
                headers=headers,
                json=payload,
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                transcribed_text = result['choices'][0]['message']['content'].strip()
                
                # Create lab note
                title = f"Lab Note - {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                lab_note = LabNote(
                    title=title,
                    content=transcribed_text,
                    project_id=project_id or 1,  # Default to first project
                    author_id=user_id
                )
                db.session.add(lab_note)
                db.session.commit()
                
                return {'success': True, 'note_id': lab_note.id, 'title': title}
            else:
                return {'success': False, 'error': f'API Error: {response.status_code} - {response.text}'}
                
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    # Route definitions
    
    @app.route('/')
    @login_required
    def index():
        """Main dashboard"""
        # Get user's projects
        user_projects = Project.query.filter_by(owner_id=current_user.id).order_by(Project.updated_at.desc()).all()
        
        # Get recent processing jobs
        recent_jobs = ProcessingJob.query.filter_by(user_id=current_user.id).order_by(ProcessingJob.created_at.desc()).limit(5).all()
        
        return render_template('index_authenticated.html', 
                             projects=user_projects, 
                             recent_jobs=recent_jobs)
    
    @app.route('/create_project', methods=['GET', 'POST'])
    @login_required
    def create_project():
        """Create new project"""
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()
            hypothesis = request.form.get('hypothesis', '').strip()
            purpose = request.form.get('purpose', '').strip()
            
            if not name:
                flash('Project name is required.', 'error')
                return render_template('create_project_authenticated.html')
            
            # Check for duplicate project names
            existing = Project.query.filter_by(owner_id=current_user.id, name=name).first()
            if existing:
                flash('A project with this name already exists.', 'error')
                return render_template('create_project_authenticated.html')
            
            try:
                project = Project(
                    name=name,
                    description=description,
                    hypothesis=hypothesis,
                    purpose=purpose,
                    owner_id=current_user.id
                )
                
                db.session.add(project)
                db.session.commit()
                
                flash(f'Project "{name}" created successfully!', 'success')
                return redirect(url_for('view_project', project_id=project.id))
                
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error creating project: {str(e)}")
                flash('Failed to create project. Please try again.', 'error')
        
        return render_template('create_project_authenticated.html')
    
    @app.route('/project/<int:project_id>')
    @login_required
    def view_project(project_id):
        """View project details and lab notes"""
        project = Project.query.get_or_404(project_id)
        
        # Check access
        if project.owner_id != current_user.id:
            flash('Access denied.', 'error')
            return redirect(url_for('index'))
        
        # Get lab notes
        lab_notes = LabNote.query.filter_by(project_id=project_id).order_by(LabNote.created_at.desc()).all()
        
        return render_template('project_authenticated.html', project=project, lab_notes=lab_notes)
    
    @app.route('/upload_file', methods=['POST'])
    @login_required
    def upload_file():
        """Upload file for processing (images, documents, etc.)"""
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Supported formats: Images (JPG, PNG, GIF, TIFF), Documents (PDF, DOC, DOCX, TXT), Spreadsheets (XLS, XLSX, CSV), Presentations (PPT, PPTX), Bioinformatics (FASTQ, FASTA, SAM, BAM, VCF, GFF, GTF, BED), Archives (ZIP, RAR).'}), 400
        
        try:
            # Save uploaded file
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            unique_filename = f"{timestamp}{uuid.uuid4().hex[:8]}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(filepath)
            
            # Determine file type
            file_type = get_file_type(filename)
            
            # Store in session for processing
            session['uploaded_file'] = filepath
            session['uploaded_file_type'] = file_type
            
            return jsonify({
                'success': True, 
                'filename': unique_filename,
                'filepath': filepath,
                'file_type': file_type,
                'original_name': filename
            })
            
        except Exception as e:
            current_app.logger.error(f"Upload error: {str(e)}")
            return jsonify({'error': 'Upload failed'}), 500
    
    # Backward compatibility route for image uploads
    @app.route('/upload_image', methods=['POST'])
    @login_required 
    def upload_image():
        """Legacy image upload endpoint for backward compatibility"""
        if 'image' not in request.files:
            return jsonify({'error': 'No image uploaded'}), 400
        
        # Convert image field to file field and reuse upload_file logic
        file = request.files['image']
        request.files = request.files.copy()
        request.files['file'] = file
        
        return upload_file()
    
    @app.route('/enhanced_ocr', methods=['POST'])
    @login_required
    def enhanced_ocr_endpoint():
        """Enhanced OCR processing with preprocessing"""
        try:
            # Check for uploaded file
            if 'uploaded_file' in session:
                file_path = session['uploaded_file']
                file_type = session.get('uploaded_file_type', 'image')
            elif 'uploaded_image' in session:
                file_path = session['uploaded_image']
                file_type = 'image'
            else:
                return jsonify({'error': 'No file uploaded'}), 400
                
            if not os.path.exists(file_path):
                return jsonify({'error': 'File not found'}), 400
                
            # Only process images
            if file_type != 'image':
                return jsonify({'error': 'Enhanced OCR is only available for image files'}), 400
            
            # Get processing options from request
            data = request.get_json() or {}
            preprocessing_type = data.get('preprocessing_type', 'auto')  # auto, standard, aggressive, handwriting
            project_id = data.get('project_id')
            
            # Initialize enhanced OCR processor
            ocr_processor = EnhancedOCR()
            
            # Process the image
            if preprocessing_type == 'auto':
                result = ocr_processor.extract_text_with_multiple_methods(file_path)
            else:
                result = ocr_processor.extract_text(file_path, preprocessing_type=preprocessing_type)
            
            # Save as lab note if project_id provided
            if project_id and result['text'].strip():
                try:
                    project = Project.query.get(project_id)
                    if project and (project.owner_id == current_user.id or 
                                  ProjectMember.query.filter_by(project_id=project_id, user_id=current_user.id).first()):
                        
                        # Create lab note
                        lab_note = LabNote(
                            filename=os.path.basename(file_path),
                            title=f"OCR Text - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                            content=result['text'],
                            processed_text=result['text'],
                            original_image_path=file_path,
                            processing_status='completed',
                            project_id=project_id,
                            author_id=current_user.id,
                            auto_classified=True
                        )
                        db.session.add(lab_note)
                        db.session.commit()
                        
                        result['lab_note_id'] = lab_note.id
                        result['saved_to_project'] = True
                        
                except Exception as e:
                    current_app.logger.error(f"Failed to save OCR result as lab note: {e}")
                    result['save_error'] = str(e)
            
            return jsonify({
                'success': True,
                'ocr_result': result,
                'message': 'Enhanced OCR processing completed successfully'
            })
            
        except Exception as e:
            current_app.logger.error(f"Enhanced OCR processing failed: {e}")
            return jsonify({'error': f'OCR processing failed: {str(e)}'}), 500

    @app.route('/process_lab_note', methods=['POST'])
    @login_required
    def process_lab_note_endpoint():
        """Process lab note image with OCR + structured parsing + database storage"""
        try:
            project_id = request.form.get('project_id')
            
            # Check for uploaded file
            if 'uploaded_file' in session:
                file_path = session['uploaded_file']
                file_type = session.get('uploaded_file_type', 'image')
            elif 'uploaded_image' in session:
                file_path = session['uploaded_image']
                file_type = 'image'
            else:
                return jsonify({'error': 'No file uploaded'}), 400
            
            if not os.path.exists(file_path):
                return jsonify({'error': 'File not found'}), 400
            
            # Only process images
            if file_type != 'image':
                return jsonify({'error': 'Lab note processing is only available for image files'}), 400
            
            # Process the lab note image
            result = process_lab_note_image(file_path, enhanced=True, db_path="lab_records.db")
            
            # Save as lab note if project_id provided
            if project_id and result['structured_data']['experiment_id']:
                try:
                    project = Project.query.get(project_id)
                    if project and (project.owner_id == current_user.id or 
                                  ProjectMember.query.filter_by(project_id=project_id, user_id=current_user.id).first()):
                        
                        structured_data = result['structured_data']
                        
                        # Create lab note with structured data
                        lab_note = LabNote(
                            filename=os.path.basename(file_path),
                            title=structured_data.get('title') or f"Lab Note - {structured_data['experiment_id']}",
                            content=result['ocr_results']['extracted_text'],
                            processed_text=json.dumps(structured_data, indent=2),
                            original_image_path=file_path,
                            processing_status='completed',
                            project_id=project_id,
                            author_id=current_user.id,
                            auto_classified=True
                        )
                        db.session.add(lab_note)
                        db.session.commit()
                        
                        result['lab_note_id'] = lab_note.id
                        result['saved_to_project'] = True
                        
                except Exception as e:
                    current_app.logger.error(f"Failed to save structured lab note: {e}")
                    result['save_error'] = str(e)
            
            return jsonify({
                'success': True,
                'result': result,
                'message': 'Lab note processed and structured successfully'
            })
            
        except Exception as e:
            current_app.logger.error(f"Lab note processing failed: {e}")
            return jsonify({'error': f'Lab note processing failed: {str(e)}'}), 500

    @app.route('/parse_text', methods=['POST'])
    @login_required
    def parse_text_endpoint():
        """Parse raw text into structured lab note format"""
        try:
            data = request.get_json()
            text = data.get('text', '')
            project_id = data.get('project_id')
            
            if not text.strip():
                return jsonify({'error': 'No text provided'}), 400
            
            # Process the text
            structured_data = process_lab_note_text(text, db_path="lab_records.db")
            
            # Save as lab note if project_id provided
            if project_id and structured_data['experiment_id']:
                try:
                    project = Project.query.get(project_id)
                    if project and (project.owner_id == current_user.id or 
                                  ProjectMember.query.filter_by(project_id=project_id, user_id=current_user.id).first()):
                        
                        # Create lab note with structured data
                        lab_note = LabNote(
                            filename="manual_text_entry.txt",
                            title=structured_data.get('title') or f"Lab Note - {structured_data['experiment_id']}",
                            content=text,
                            processed_text=json.dumps(structured_data, indent=2),
                            processing_status='completed',
                            project_id=project_id,
                            author_id=current_user.id,
                            auto_classified=True
                        )
                        db.session.add(lab_note)
                        db.session.commit()
                        
                        structured_data['lab_note_id'] = lab_note.id
                        structured_data['saved_to_project'] = True
                        
                except Exception as e:
                    current_app.logger.error(f"Failed to save parsed text as lab note: {e}")
                    structured_data['save_error'] = str(e)
            
            return jsonify({
                'success': True,
                'structured_data': structured_data,
                'message': 'Text parsed and structured successfully'
            })
            
        except Exception as e:
            current_app.logger.error(f"Text parsing failed: {e}")
            return jsonify({'error': f'Text parsing failed: {str(e)}'}), 500

    @app.route('/lab_records', methods=['GET'])
    @login_required
    def get_lab_records_endpoint():
        """Get structured lab records from database"""
        try:
            limit = request.args.get('limit', 50, type=int)
            
            parser = LabNoteParser(db_path="lab_records.db")
            records = parser.list_experiments(limit=limit)
            
            return jsonify({
                'success': True,
                'records': records,
                'count': len(records)
            })
            
        except Exception as e:
            current_app.logger.error(f"Failed to retrieve lab records: {e}")
            return jsonify({'error': f'Failed to retrieve lab records: {str(e)}'}), 500

    @app.route('/lab_records/<experiment_id>', methods=['GET'])
    @login_required
    def get_lab_record_endpoint(experiment_id):
        """Get specific lab record by experiment ID"""
        try:
            parser = LabNoteParser(db_path="lab_records.db")
            record = parser.get_experiment(experiment_id)
            
            if record:
                return jsonify({
                    'success': True,
                    'record': record
                })
            else:
                return jsonify({'error': 'Lab record not found'}), 404
            
        except Exception as e:
            current_app.logger.error(f"Failed to retrieve lab record {experiment_id}: {e}")
            return jsonify({'error': f'Failed to retrieve lab record: {str(e)}'}), 500

    @app.route('/search_lab_records', methods=['GET'])
    @login_required
    def search_lab_records_endpoint():
        """Search lab records using full-text search"""
        try:
            query = request.args.get('q', '').strip()
            limit = request.args.get('limit', 50, type=int)
            
            if not query:
                return jsonify({'error': 'Search query is required'}), 400
            
            parser = LabNoteParser(db_path="lab_records.db")
            results = parser.search_experiments(query, limit=limit)
            
            return jsonify({
                'success': True,
                'results': results,
                'query': query,
                'count': len(results)
            })
            
        except Exception as e:
            current_app.logger.error(f"Search failed: {e}")
            return jsonify({'error': f'Search failed: {str(e)}'}), 500

    @app.route('/search_measurements', methods=['GET'])
    @login_required
    def search_measurements_endpoint():
        """Search measurements by type and/or value range"""
        try:
            measurement_type = request.args.get('type')
            min_value = request.args.get('min_value', type=float)
            max_value = request.args.get('max_value', type=float)
            limit = request.args.get('limit', 50, type=int)
            
            parser = LabNoteParser(db_path="lab_records.db")
            results = parser.search_measurements(
                measurement_type=measurement_type,
                min_value=min_value,
                max_value=max_value,
                limit=limit
            )
            
            return jsonify({
                'success': True,
                'results': results,
                'filters': {
                    'type': measurement_type,
                    'min_value': min_value,
                    'max_value': max_value
                },
                'count': len(results)
            })
            
        except Exception as e:
            current_app.logger.error(f"Measurement search failed: {e}")
            return jsonify({'error': f'Measurement search failed: {str(e)}'}), 500

    @app.route('/search_suggestions', methods=['GET'])
    @login_required
    def search_suggestions_endpoint():
        """Get search suggestions for autocomplete"""
        try:
            partial_query = request.args.get('q', '').strip()
            limit = request.args.get('limit', 10, type=int)
            
            if len(partial_query) < 2:
                return jsonify({
                    'success': True,
                    'suggestions': {
                        'experiment_ids': [],
                        'researchers': [],
                        'titles': [],
                        'measurement_types': []
                    }
                })
            
            parser = LabNoteParser(db_path="lab_records.db")
            suggestions = parser.get_search_suggestions(partial_query, limit=limit)
            
            return jsonify({
                'success': True,
                'suggestions': suggestions,
                'query': partial_query
            })
            
        except Exception as e:
            current_app.logger.error(f"Search suggestions failed: {e}")
            return jsonify({'error': f'Search suggestions failed: {str(e)}'}), 500

    @app.route('/process_image_async', methods=['POST'])
    @login_required
    def process_image_endpoint():
        """Start async image processing"""
        data = request.get_json()
        api_key = data.get('api_key')
        project_id = data.get('project_id')  # Optional
        
        if not api_key:
            return jsonify({'error': 'xAI API key required'}), 400
        
        # Check for uploaded file (new format) or image (legacy)
        if 'uploaded_file' in session:
            file_path = session['uploaded_file']
            file_type = session.get('uploaded_file_type', 'image')
        elif 'uploaded_image' in session:
            file_path = session['uploaded_image']
            file_type = 'image'
        else:
            return jsonify({'error': 'No file uploaded'}), 400
            
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 400
            
        # Only process images with OCR for now
        if file_type != 'image':
            return jsonify({'error': 'OCR processing is only available for image files. Other file types will be stored as project documents.'}), 400
        
        try:
            # Create processing job record
            job_id = str(uuid.uuid4())
            processing_job = ProcessingJob(
                id=job_id,
                job_type='image_processing',
                user_id=current_user.id,
                project_id=project_id
            )
            db.session.add(processing_job)
            db.session.commit()
            
            # Start async task
            task = process_image_async.apply_async(
                args=[file_path, api_key, current_user.id, project_id],
                task_id=job_id
            )
            
            return jsonify({
                'success': True,
                'job_id': job_id,
                'message': 'Image processing started'
            })
            
        except Exception as e:
            current_app.logger.error(f"Process start error: {str(e)}")
            return jsonify({'error': 'Failed to start processing'}), 500
    
    @app.route('/api/upload', methods=['POST'])
    @jwt_required()
    def api_upload_file():
        """General file upload endpoint with JWT authentication"""
        user_id = int(get_jwt_identity())
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Supported formats: Images (JPG, PNG, GIF, TIFF), Documents (PDF, DOC, DOCX, TXT), Spreadsheets (XLS, XLSX, CSV), Presentations (PPT, PPTX), Bioinformatics (FASTQ, FASTA, SAM, BAM, VCF, GFF, GTF, BED), Archives (ZIP, RAR).'}), 400
        
        try:
            # Save uploaded file
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            unique_filename = f"{timestamp}{uuid.uuid4().hex[:8]}_{filename}"
            
            # Save to uploads folder
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
            file.save(filepath)
            
            # For images, automatically start AI processing
            is_image = False
            if file.content_type and file.content_type.startswith('image/'):
                is_image = True
            elif filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff')):
                is_image = True
            
            if is_image:
                # Store in session for processing
                session['uploaded_file'] = filepath
                session['uploaded_file_type'] = 'image'
                
                # Get API key from form data
                api_key = request.form.get('api_key')
                if not api_key:
                    return jsonify({
                        'success': True,
                        'message': 'Image uploaded successfully. Please set your xAI API key in Settings to enable AI processing.',
                        'filename': unique_filename,
                        'filepath': filepath,
                        'file_type': 'image',
                        'needs_processing': True,
                        'processing_started': False
                    })
                
                project_id = None  # Will auto-classify to appropriate project
                
                try:
                    # Process image directly (instead of using Celery for now)
                    result = process_image_direct(filepath, api_key, user_id, project_id)
                    
                    if result['success']:
                        return jsonify({
                            'success': True,
                            'message': 'Image uploaded and processed successfully',
                            'filename': unique_filename,
                            'filepath': filepath,
                            'file_type': 'image',
                            'note_id': result['note_id'],
                            'note_title': result['title'],
                            'processing_completed': True
                        })
                    else:
                        return jsonify({
                            'success': False,
                            'error': result.get('error', 'Processing failed'),
                            'filename': unique_filename,
                            'filepath': filepath,
                            'file_type': 'image'
                        }), 500
                    
                except Exception as e:
                    # If processing fails, still return success for upload
                    return jsonify({
                        'success': True,
                        'message': 'Image uploaded successfully (processing failed to start)',
                        'filename': unique_filename,
                        'filepath': filepath,  
                        'file_type': 'image',
                        'processing_error': str(e)
                    })
            else:
                return jsonify({
                    'success': True,
                    'message': 'File uploaded successfully',
                    'filename': unique_filename,
                    'filepath': filepath,
                    'file_type': 'document'
                })
            
        except Exception as e:
            return jsonify({'error': f'Upload failed: {str(e)}'}), 500

    @app.route('/api/projects/<int:project_id>/upload', methods=['POST'])
    @jwt_required()
    def upload_project_file(project_id):
        """Upload file directly to project"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project
        if project.owner_id != user_id:
            # TODO: Check if user is a member with edit permissions
            return jsonify({'error': 'Access denied'}), 403
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Supported formats: Images (JPG, PNG, GIF, TIFF), Documents (PDF, DOC, DOCX, TXT), Spreadsheets (XLS, XLSX, CSV), Presentations (PPT, PPTX), Bioinformatics (FASTQ, FASTA, SAM, BAM, VCF, GFF, GTF, BED), Archives (ZIP, RAR).'}), 400
        
        try:
            # Save uploaded file
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
            unique_filename = f"{timestamp}{uuid.uuid4().hex[:8]}_{filename}"
            
            # Create project-specific folder
            project_folder = os.path.join(app.config['PROJECTS_FOLDER'], str(project_id))
            os.makedirs(project_folder, exist_ok=True)
            
            filepath = os.path.join(project_folder, unique_filename)
            file.save(filepath)
            
            # Determine file type
            file_type = get_file_type(filename)
            
            # Create a lab note entry for the file
            lab_note = LabNote(
                title=f"Uploaded: {filename}",
                content=f"File upload: {filename} ({file_type})\nFile type: {file_type}\nOriginal filename: {filename}\nStored as: {unique_filename}",
                original_image_path=filepath,
                processing_status='completed',  # No processing needed for non-image files
                project_id=project_id,
                author_id=user_id
            )
            
            db.session.add(lab_note)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'lab_note_id': lab_note.id,
                'filename': unique_filename,
                'original_name': filename,
                'file_type': file_type,
                'message': f'{file_type.title()} file uploaded successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Project file upload error: {str(e)}")
            return jsonify({'error': 'Upload failed'}), 500
    
    @app.route('/job_status/<job_id>')
    @login_required
    def job_status(job_id):
        """Get processing job status"""
        job = ProcessingJob.query.get_or_404(job_id)
        
        if job.user_id != current_user.id:
            return jsonify({'error': 'Access denied'}), 403
        
        # Check Celery task status
        task = app.celery.AsyncResult(job_id)
        
        response_data = {
            'job_id': job_id,
            'status': job.status,
            'progress': job.progress,
            'error_message': job.error_message
        }
        
        if job.status == 'completed' and job.result:
            try:
                response_data['result'] = json.loads(job.result)
            except:
                response_data['result'] = job.result
        
        return jsonify(response_data)
    
    @app.route('/generate_report_async/<int:project_id>', methods=['POST'])
    @login_required
    def generate_report_async_endpoint(project_id):
        """Start async report generation"""
        data = request.get_json()
        api_key = data.get('api_key')
        
        if not api_key:
            return jsonify({'error': 'xAI API key required'}), 400
        
        project = Project.query.get_or_404(project_id)
        if project.owner_id != current_user.id:
            return jsonify({'error': 'Access denied'}), 403
        
        try:
            # Create processing job
            job_id = str(uuid.uuid4())
            processing_job = ProcessingJob(
                id=job_id,
                job_type='report_generation',
                user_id=current_user.id,
                project_id=project_id
            )
            db.session.add(processing_job)
            db.session.commit()
            
            # Start async task
            task = generate_project_report_async.apply_async(
                args=[project_id, api_key, current_user.id],
                task_id=job_id
            )
            
            return jsonify({
                'success': True,
                'job_id': job_id,
                'message': 'Report generation started'
            })
            
        except Exception as e:
            current_app.logger.error(f"Report generation error: {str(e)}")
            return jsonify({'error': 'Failed to start report generation'}), 500
    
    @app.route('/api/projects/legacy')
    @login_required
    def api_projects_legacy():
        """Get user's projects (Legacy Flask-Login API endpoint)"""
        projects = Project.query.filter_by(owner_id=current_user.id).order_by(Project.updated_at.desc()).all()
        return jsonify({
            'projects': [project.to_dict() for project in projects]
        })
    
    @app.route('/api/lab_notes/<int:project_id>')
    @login_required
    def api_lab_notes(project_id):
        """Get lab notes for a project (API endpoint)"""
        project = Project.query.get_or_404(project_id)
        if project.owner_id != current_user.id:
            return jsonify({'error': 'Access denied'}), 403
        
        lab_notes = LabNote.query.filter_by(project_id=project_id).order_by(LabNote.created_at.desc()).all()
        return jsonify({
            'lab_notes': [note.to_dict() for note in lab_notes]
        })
    
    @app.route('/api/notes/<int:project_id>')
    @jwt_required()
    def api_notes_jwt(project_id):
        """Get lab notes for a project via JWT"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project
        is_owner = project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        lab_notes = LabNote.query.filter_by(project_id=project_id).order_by(LabNote.created_at.desc()).all()
        
        # Convert lab notes to dict with additional metadata
        notes_data = []
        for note in lab_notes:
            note_dict = note.to_dict()
            # Add additional fields for frontend compatibility
            note_dict.update({
                'filename': note.title,
                'file_size': 1024,  # Default file size if not available
                'processed_text': note.content,
                'processing_status': note.processing_status or 'completed'
            })
            notes_data.append(note_dict)
        
        return jsonify({
            'notes': notes_data
        })
    
    @app.route('/api/notes', methods=['GET'])
    @jwt_required()
    def api_notes_all_jwt():
        """Get all lab notes for user or filter by project_id via JWT"""
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Check if filtering by project_id
        project_id = request.args.get('project_id')
        
        if project_id:
            try:
                project_id = int(project_id)
                project = Project.query.get_or_404(project_id)
                
                # Check if user has access to this project
                is_owner = project.owner_id == user_id
                is_member = False
                
                if not is_owner:
                    member_check = ProjectMember.query.filter_by(
                        project_id=project_id,
                        user_id=user_id
                    ).first()
                    is_member = member_check is not None
                
                if not (is_owner or is_member):
                    return jsonify({'error': 'Access denied'}), 403
                
                lab_notes = LabNote.query.filter_by(project_id=project_id).order_by(LabNote.created_at.desc()).all()
            except ValueError:
                return jsonify({'error': 'Invalid project_id'}), 400
        else:
            # Get all notes for projects the user owns or is a member of
            owned_projects = Project.query.filter_by(owner_id=user_id).all()
            member_projects = db.session.query(Project).join(ProjectMember).filter(
                ProjectMember.user_id == user_id
            ).all()
            
            all_project_ids = [p.id for p in owned_projects] + [p.id for p in member_projects]
            
            if all_project_ids:
                lab_notes = LabNote.query.filter(
                    LabNote.project_id.in_(all_project_ids)
                ).order_by(LabNote.created_at.desc()).all()
            else:
                lab_notes = []
        
        # Convert lab notes to dict with additional metadata
        notes_data = []
        for note in lab_notes:
            note_dict = note.to_dict()
            # Add additional fields for frontend compatibility
            note_dict.update({
                'filename': note.title,
                'file_size': 1024,  # Default file size if not available
                'processed_text': note.content,
                'processing_status': note.processing_status or 'completed'
            })
            notes_data.append(note_dict)
        
        return jsonify({
            'notes': notes_data
        })
    
    @app.route('/api/notes/<int:note_id>', methods=['DELETE'])
    @jwt_required()
    def api_delete_note(note_id):
        """Delete a specific lab note"""
        user_id = int(get_jwt_identity())
        note = LabNote.query.get_or_404(note_id)
        
        # Check if user has permission to delete this note
        # User can delete if they are the author or owner of the project
        is_author = note.author_id == user_id
        is_project_owner = note.project.owner_id == user_id
        
        # Check if user is project admin/member with delete permissions
        is_project_admin = False
        if not is_author and not is_project_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=note.project_id,
                user_id=user_id,
                role='admin'
            ).first()
            is_project_admin = member_check is not None
        
        if not (is_author or is_project_owner or is_project_admin):
            return jsonify({'error': 'Permission denied'}), 403
        
        try:
            # Delete associated files if they exist
            if note.original_image_path:
                import os
                file_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), note.original_image_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
            
            db.session.delete(note)
            db.session.commit()
            
            return jsonify({'success': True, 'message': 'Note deleted successfully'})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to delete note: {str(e)}'}), 500
    
    # Dashboard API endpoints
    @app.route('/api/dashboard/stats', methods=['GET'])
    @jwt_required()
    def api_dashboard_stats():
        """Get dashboard statistics for the authenticated user"""
        user_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        try:
            # Get user's projects (owned + member projects)
            owned_projects = Project.query.filter_by(owner_id=user_id)
            member_project_ids = db.session.query(ProjectMember.project_id).filter_by(user_id=user_id)
            member_projects = Project.query.filter(Project.id.in_(member_project_ids))
            all_projects = owned_projects.union(member_projects).all()
            
            # Get lab notes count for all accessible projects
            if all_projects:
                project_ids = [p.id for p in all_projects]
                notes_count = LabNote.query.filter(LabNote.project_id.in_(project_ids)).count()
            else:
                notes_count = 0
            
            # Get processing jobs count
            processing_count = ProcessingJob.query.filter_by(user_id=user_id, status='processing').count()
            
            return jsonify({
                'stats': {
                    'total_notes': notes_count,
                    'projects': len(all_projects),
                    'processing': processing_count
                }
            })
            
        except Exception as e:
            current_app.logger.error(f"Dashboard stats error: {str(e)}")
            return jsonify({'error': 'Failed to fetch dashboard stats'}), 500

    # JWT API endpoints for frontend
    
    @app.route('/api/projects', methods=['GET', 'POST'])
    @jwt_required()
    def api_projects_jwt():
        """Get user's projects via JWT or create new project"""
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if request.method == 'POST':
            # Create new project
            data = request.get_json()
            
            if not data or not data.get('name'):
                return jsonify({'error': 'Project name is required'}), 400
            
            name = data['name'].strip()
            description = data.get('description', '').strip()
            hypothesis = data.get('hypothesis', '').strip()
            purpose = data.get('purpose', '').strip()
            future_plan = data.get('future_plan', '').strip()
            
            # Check for existing project
            existing = Project.query.filter_by(owner_id=user_id, name=name).first()
            if existing:
                return jsonify({'error': 'A project with this name already exists'}), 400
            
            try:
                project = Project(
                    name=name,
                    description=description,
                    hypothesis=hypothesis,
                    purpose=purpose,
                    future_plan=future_plan,
                    owner_id=user_id
                )
                
                db.session.add(project)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'project': project.to_dict()
                })
                
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error creating project: {str(e)}")
                return jsonify({'error': 'Failed to create project'}), 500
        
        # GET request - return projects (owned + member projects)
        from sqlalchemy import or_
        
        # Get projects where user is owner or member
        owned_projects = Project.query.filter_by(owner_id=user_id)
        member_project_ids = db.session.query(ProjectMember.project_id).filter_by(user_id=user_id)
        member_projects = Project.query.filter(Project.id.in_(member_project_ids))
        
        # Combine and order by updated_at
        all_projects = owned_projects.union(member_projects).order_by(Project.updated_at.desc()).all()
        
        projects_data = []
        for project in all_projects:
            project_dict = project.to_dict()
            
            # Add user-specific information
            is_owner = project.owner_id == user_id
            user_role = 'owner' if is_owner else None
            
            if not is_owner:
                # Get user's role as member
                member = ProjectMember.query.filter_by(
                    project_id=project.id,
                    user_id=user_id
                ).first()
                if member:
                    user_role = member.role
            
            project_dict['is_owner'] = is_owner
            project_dict['user_role'] = user_role
            
            # Add member count
            member_count = ProjectMember.query.filter_by(project_id=project.id).count()
            project_dict['member_count'] = member_count + 1  # +1 for owner
            
            projects_data.append(project_dict)
        
        return jsonify({
            'projects': projects_data
        })
    
    @app.route('/api/projects/<int:project_id>', methods=['PUT'])
    @jwt_required()
    def api_update_project(project_id):
        """Update project details"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project
        is_owner = project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None and member_check.role in ['admin', 'member']
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        try:
            # Update allowed fields
            if 'name' in data and is_owner:  # Only owner can change name
                project.name = data['name'].strip()
            if 'description' in data:
                project.description = data['description'].strip()
            if 'hypothesis' in data:
                project.hypothesis = data['hypothesis'].strip()
            if 'purpose' in data:
                project.purpose = data['purpose'].strip()
            if 'future_plan' in data:
                project.future_plan = data['future_plan'].strip()
            if 'current_progress' in data:
                project.current_progress = data['current_progress'].strip()
            if 'progress_percentage' in data:
                progress = int(data['progress_percentage'])
                if 0 <= progress <= 100:
                    project.progress_percentage = progress
            if 'status' in data and is_owner:  # Only owner can change status
                if data['status'] in ['active', 'completed', 'paused']:
                    project.status = data['status']
            
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Project updated successfully',
                'project': project.to_dict()
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'Failed to update project'}), 500
    
    # Project Collaboration APIs
    @app.route('/api/projects/<int:project_id>/members', methods=['GET'])
    @jwt_required()
    def api_project_members(project_id):
        """Get project members"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project (owner or member)
        is_owner = project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        # Get all members
        members = []
        
        # Add owner first
        members.append({
            'id': None,
            'user_id': project.owner_id,
            'username': project.owner.username,
            'email': project.owner.email,
            'full_name': project.owner.get_full_name(),
            'role': 'owner',
            'joined_at': project.created_at.isoformat(),
            'is_owner': True,
            'can_edit': False,
            'can_remove': False
        })
        
        # Add all project members
        project_members = ProjectMember.query.filter_by(project_id=project_id).all()
        for member in project_members:
            members.append({
                'id': member.id,
                'user_id': member.user_id,
                'username': member.user.username,
                'email': member.user.email,
                'full_name': member.user.get_full_name(),
                'role': member.role,
                'joined_at': member.joined_at.isoformat(),
                'is_owner': False,
                'can_edit': member.role in ['admin', 'member'],
                'can_remove': is_owner  # Only owner can remove members
            })
        
        return jsonify({'members': members})
    
    @app.route('/api/projects/<int:project_id>/members', methods=['POST'])
    @jwt_required()
    def api_add_project_member(project_id):
        """Add member to project"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        if project.owner_id != user_id:
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data or not data.get('email') or not data.get('role'):
            return jsonify({'error': 'Email and role are required'}), 400
        
        email = data['email'].strip().lower()
        role = data['role']
        
        if role not in ['admin', 'member', 'viewer']:
            return jsonify({'error': 'Invalid role. Must be admin, member, or viewer'}), 400
        
        # Find user by email
        target_user = User.query.filter_by(email=email).first()
        if not target_user:
            return jsonify({'error': f'User with email {email} not found'}), 404
        
        # Check if user is already a member
        existing_member = ProjectMember.query.filter_by(
            project_id=project_id,
            user_id=target_user.id
        ).first()
        
        if existing_member:
            return jsonify({'error': 'User is already a member of this project'}), 400
        
        # Don't allow adding the owner as a member
        if target_user.id == project.owner_id:
            return jsonify({'error': 'Project owner cannot be added as a member'}), 400
        
        try:
            # Create new project member
            new_member = ProjectMember(
                project_id=project_id,
                user_id=target_user.id,
                role=role
            )
            
            db.session.add(new_member)
            db.session.commit()
            
            # Send email notification to invited user
            inviter = User.query.get(user_id)
            email_template = """
            <html>
            <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
                <div style="background-color: #f8f9fa; padding: 20px; border-radius: 8px;">
                    <h2 style="color: #2563eb; margin-bottom: 20px;">Project Invitation</h2>
                    
                    <p>Hello {user_name},</p>
                    
                    <p>You have been invited to collaborate on the project <strong>"{project_name}"</strong> by {inviter_name}.</p>
                    
                    <div style="background-color: white; padding: 15px; border-radius: 5px; margin: 20px 0;">
                        <h3 style="margin-top: 0; color: #374151;">Project Details:</h3>
                        <p><strong>Name:</strong> {project_name}</p>
                        <p><strong>Your Role:</strong> {role}</p>
                        <p><strong>Invited by:</strong> {inviter_name}</p>
                    </div>
                    
                    <p>You can now access this project and its lab notes in your dashboard.</p>
                    
                    <div style="margin-top: 30px;">
                        <a href="http://10.67.36.36:3000/dashboard" 
                           style="background-color: #2563eb; color: white; padding: 12px 24px; 
                                  text-decoration: none; border-radius: 5px; display: inline-block;">
                            Go to Dashboard
                        </a>
                    </div>
                    
                    <p style="margin-top: 30px; color: #6b7280; font-size: 14px;">
                        This is an automated message from Lab Notebook.
                    </p>
                </div>
            </body>
            </html>
            """
            
            send_email(
                to=target_user.email,
                subject=f"You've been invited to collaborate on '{project.name}'",
                template=email_template,
                user_name=target_user.get_full_name(),
                project_name=project.name,
                role=role.title(),
                inviter_name=inviter.get_full_name()
            )
            
            return jsonify({
                'success': True,
                'member': {
                    'id': new_member.id,
                    'user_id': target_user.id,
                    'username': target_user.username,
                    'email': target_user.email,
                    'full_name': target_user.get_full_name(),
                    'role': role,
                    'joined_at': new_member.joined_at.isoformat(),
                    'is_owner': False,
                    'can_edit': role in ['admin', 'member'],
                    'can_remove': True
                }
            })
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error adding project member: {str(e)}")
            return jsonify({'error': 'Failed to add member'}), 500
    
    @app.route('/api/projects/<int:project_id>/members/<int:member_id>', methods=['PUT'])
    @jwt_required()
    def api_update_project_member(project_id, member_id):
        """Update member role"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        if project.owner_id != user_id:
            return jsonify({'error': 'Access denied'}), 403
        
        member = ProjectMember.query.filter_by(
            id=member_id,
            project_id=project_id
        ).first()
        
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        data = request.get_json()
        if not data or not data.get('role'):
            return jsonify({'error': 'Role is required'}), 400
        
        new_role = data['role']
        if new_role not in ['admin', 'member', 'viewer']:
            return jsonify({'error': 'Invalid role. Must be admin, member, or viewer'}), 400
        
        try:
            member.role = new_role
            db.session.commit()
            
            return jsonify({
                'success': True,
                'member': {
                    'id': member.id,
                    'user_id': member.user_id,
                    'username': member.user.username,
                    'email': member.user.email,
                    'full_name': member.user.get_full_name(),
                    'role': new_role,
                    'joined_at': member.joined_at.isoformat(),
                    'is_owner': False,
                    'can_edit': new_role in ['admin', 'member'],
                    'can_remove': True
                }
            })
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating project member: {str(e)}")
            return jsonify({'error': 'Failed to update member role'}), 500
    
    @app.route('/api/projects/<int:project_id>/members/<int:member_id>', methods=['DELETE'])
    @jwt_required()
    def api_remove_project_member(project_id, member_id):
        """Remove member from project"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        if project.owner_id != user_id:
            return jsonify({'error': 'Access denied'}), 403
        
        member = ProjectMember.query.filter_by(
            id=member_id,
            project_id=project_id
        ).first()
        
        if not member:
            return jsonify({'error': 'Member not found'}), 404
        
        try:
            db.session.delete(member)
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error removing project member: {str(e)}")
            return jsonify({'error': 'Failed to remove member'}), 500
    
    @app.route('/api/projects/<int:project_id>/export/excel', methods=['GET'])
    @jwt_required()
    def export_project_excel(project_id):
        """Export project data and lab notes to Excel"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project
        is_owner = project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        try:
            # Create Excel workbook
            wb = Workbook()
            
            # Project Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Project Summary"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Project information
            ws_summary['A1'] = "Project Information"
            ws_summary['A1'].font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:B1')
            
            project_data = [
                ["Project Name", project.name],
                ["Description", project.description or "No description"],
                ["Hypothesis", project.hypothesis or "No hypothesis"],
                ["Purpose", project.purpose or "No purpose"],
                ["Owner", project.owner.get_full_name()],
                ["Status", project.status],
                ["Created", project.created_at.strftime('%Y-%m-%d %H:%M:%S')],
                ["Last Updated", project.updated_at.strftime('%Y-%m-%d %H:%M:%S')],
                ["Total Notes", project.get_note_count()]
            ]
            
            for i, (label, value) in enumerate(project_data, start=2):
                ws_summary[f'A{i}'] = label
                ws_summary[f'B{i}'] = value
                ws_summary[f'A{i}'].font = Font(bold=True)
            
            # Lab Notes Sheet
            lab_notes = LabNote.query.filter_by(project_id=project_id).order_by(LabNote.created_at.desc()).all()
            
            if lab_notes:
                ws_notes = wb.create_sheet("Lab Notes")
                
                # Headers
                headers = ['ID', 'Title', 'Image Path', 'Processing Status', 
                          'Created Date', 'Author', 'Processed Text Preview']
                
                for col, header in enumerate(headers, 1):
                    cell = ws_notes.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Data rows
                for row, note in enumerate(lab_notes, 2):
                    ws_notes.cell(row=row, column=1, value=note.id)
                    ws_notes.cell(row=row, column=2, value=note.title)
                    ws_notes.cell(row=row, column=3, value=note.original_image_path or 'No image')
                    ws_notes.cell(row=row, column=4, value=note.processing_status.capitalize())
                    ws_notes.cell(row=row, column=5, value=note.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws_notes.cell(row=row, column=6, value=note.author.get_full_name() if note.author else 'Unknown')
                    
                    # Processed text preview (first 100 characters)
                    preview = note.content[:100] + "..." if note.content and len(note.content) > 100 else (note.content or "No processed text")
                    ws_notes.cell(row=row, column=7, value=preview)
                
                # Auto-adjust column widths
                for column in ws_notes.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws_notes.column_dimensions[column_letter].width = adjusted_width
            
            # Project Members Sheet
            members = []
            # Add owner
            members.append({
                'Name': project.owner.get_full_name(),
                'Username': project.owner.username,
                'Email': project.owner.email,
                'Role': 'Owner',
                'Joined Date': project.created_at.strftime('%Y-%m-%d %H:%M:%S')
            })
            
            # Add project members
            project_members = ProjectMember.query.filter_by(project_id=project_id).all()
            for member in project_members:
                members.append({
                    'Name': member.user.get_full_name(),
                    'Username': member.user.username,
                    'Email': member.user.email,
                    'Role': member.role.capitalize(),
                    'Joined Date': member.joined_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            if members:
                ws_members = wb.create_sheet("Project Members")
                
                # Headers
                member_headers = ['Name', 'Username', 'Email', 'Role', 'Joined Date']
                for col, header in enumerate(member_headers, 1):
                    cell = ws_members.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                
                # Data rows
                for row, member in enumerate(members, 2):
                    ws_members.cell(row=row, column=1, value=member['Name'])
                    ws_members.cell(row=row, column=2, value=member['Username'])
                    ws_members.cell(row=row, column=3, value=member['Email'])
                    ws_members.cell(row=row, column=4, value=member['Role'])
                    ws_members.cell(row=row, column=5, value=member['Joined Date'])
                
                # Auto-adjust column widths
                for column in ws_members.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 40)
                    ws_members.column_dimensions[column_letter].width = adjusted_width
            
            # Save to memory
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{project.name}_{timestamp}_export.xlsx".replace(' ', '_')
            
            return send_file(
                excel_buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            current_app.logger.error(f"Excel export error: {str(e)}")
            return jsonify({'error': 'Export failed'}), 500
    
    # Protocol Management APIs
    @app.route('/api/projects/<int:project_id>/protocols', methods=['GET', 'POST'])
    @jwt_required()
    def api_project_protocols(project_id):
        """Get or create protocols for a project"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project
        is_owner = project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        if request.method == 'POST':
            # Create new protocol
            data = request.get_json()
            
            if not data or not data.get('name') or not data.get('content'):
                return jsonify({'error': 'Protocol name and content are required'}), 400
            
            try:
                protocol = Protocol(
                    name=data['name'].strip(),
                    content=data['content'].strip(),
                    project_id=project_id,
                    author_id=user_id
                )
                
                db.session.add(protocol)
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'protocol': protocol.to_dict()
                })
                
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error creating protocol: {str(e)}")
                return jsonify({'error': 'Failed to create protocol'}), 500
        
        # GET request - return protocols
        protocols = Protocol.query.filter_by(project_id=project_id).order_by(Protocol.updated_at.desc()).all()
        return jsonify({
            'protocols': [protocol.to_dict() for protocol in protocols]
        })
    
    @app.route('/api/protocols/<int:protocol_id>', methods=['GET', 'PUT', 'DELETE'])
    @jwt_required()
    def api_protocol_detail(protocol_id):
        """Get, update, or delete a specific protocol"""
        user_id = int(get_jwt_identity())
        protocol = Protocol.query.get_or_404(protocol_id)
        
        # Check if user has access to this protocol's project
        is_owner = protocol.project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=protocol.project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        if request.method == 'GET':
            return jsonify({'protocol': protocol.to_dict()})
        
        elif request.method == 'PUT':
            # Update protocol
            data = request.get_json()
            
            if not data:
                return jsonify({'error': 'No data provided'}), 400
            
            try:
                old_content = protocol.content
                
                if 'name' in data:
                    protocol.name = data['name'].strip()
                
                if 'content' in data:
                    new_content = data['content'].strip()
                    if new_content != old_content:
                        # Add change to history
                        change_description = data.get('change_description', 'Manual update')
                        lab_note_id = data.get('lab_note_id', None)
                        protocol.add_change(old_content, new_content, change_description, lab_note_id)
                
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'protocol': protocol.to_dict()
                })
                
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error updating protocol: {str(e)}")
                return jsonify({'error': 'Failed to update protocol'}), 500
        
        elif request.method == 'DELETE':
            if not is_owner:
                return jsonify({'error': 'Only project owner can delete protocols'}), 403
            
            try:
                db.session.delete(protocol)
                db.session.commit()
                
                return jsonify({'success': True})
                
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error deleting protocol: {str(e)}")
                return jsonify({'error': 'Failed to delete protocol'}), 500
    
    @app.route('/api/projects/<int:project_id>/progress', methods=['PUT'])
    @jwt_required()
    def api_update_project_progress(project_id):
        """Update project progress percentage"""
        user_id = int(get_jwt_identity())
        project = Project.query.get_or_404(project_id)
        
        # Check if user has access to this project
        is_owner = project.owner_id == user_id
        is_member = False
        
        if not is_owner:
            member_check = ProjectMember.query.filter_by(
                project_id=project_id,
                user_id=user_id
            ).first()
            is_member = member_check is not None and member_check.role in ['admin', 'member']
        
        if not (is_owner or is_member):
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data or 'progress_percentage' not in data:
            return jsonify({'error': 'Progress percentage is required'}), 400
        
        progress = data['progress_percentage']
        if not isinstance(progress, int) or progress < 0 or progress > 100:
            return jsonify({'error': 'Progress must be an integer between 0 and 100'}), 400
        
        try:
            project.progress_percentage = progress
            if progress == 100 and project.status != 'completed':
                project.status = 'completed'
            elif progress == 0 and project.status == 'completed':
                project.status = 'active'
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'project': project.to_dict()
            })
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating project progress: {str(e)}")
            return jsonify({'error': 'Failed to update progress'}), 500
    
    # User Search API
    @app.route('/api/users/search', methods=['GET'])
    @jwt_required()
    def api_search_users():
        """Search users for project invitations"""
        user_id = int(get_jwt_identity())
        query = request.args.get('q', '').strip()
        
        if len(query) < 1:
            return jsonify({'users': [], 'total_users': 0})
        
        try:
            # Search users by name, username, or email
            users_query = User.query.filter(
                db.or_(
                    User.username.ilike(f'%{query}%'),
                    User.email.ilike(f'%{query}%'),
                    User.first_name.ilike(f'%{query}%'),
                    User.last_name.ilike(f'%{query}%')
                ),
                User.is_active == True,
                User.id != user_id  # Exclude current user
            ).limit(10)
            
            users = users_query.all()
            
            # Get total registered users count
            total_users = User.query.filter(User.is_active == True).count()
            
            user_list = []
            for user in users:
                user_list.append({
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'full_name': user.get_full_name(),
                    'first_name': user.first_name,
                    'last_name': user.last_name
                })
            
            return jsonify({
                'users': user_list,
                'total_users': total_users,
                'query': query
            })
            
        except Exception as e:
            current_app.logger.error(f"User search error: {str(e)}")
            return jsonify({'error': 'Search failed', 'users': [], 'total_users': 0}), 500
    
    @app.route('/api/users/available', methods=['GET'])
    @jwt_required()
    def api_available_users():
        """Get all available users for project invitations"""
        user_id = int(get_jwt_identity())
        project_id = request.args.get('project_id', type=int)
        
        try:
            # Get all active users except current user
            users_query = User.query.filter(
                User.is_active == True,
                User.id != user_id
            )
            
            # If project_id is provided, exclude users who are already members
            if project_id:
                # Get existing project members
                existing_members = db.session.query(ProjectMember.user_id).filter_by(project_id=project_id).subquery()
                project = Project.query.get(project_id)
                
                # Exclude project owner and existing members
                users_query = users_query.filter(
                    User.id != project.owner_id if project else True,
                    ~User.id.in_(existing_members)
                )
            
            users = users_query.order_by(User.first_name, User.last_name).all()
            
            # Get total registered users count
            total_users = User.query.filter(User.is_active == True).count()
            
            user_list = []
            for user in users:
                user_list.append({
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'full_name': user.get_full_name(),
                    'first_name': user.first_name,
                    'last_name': user.last_name
                })
            
            return jsonify({
                'users': user_list,
                'total_users': total_users,
                'available_count': len(user_list)
            })
            
        except Exception as e:
            current_app.logger.error(f"Available users error: {str(e)}")
            return jsonify({'error': 'Failed to fetch users', 'users': [], 'total_users': 0}), 500
    
    # Helper functions
    def allowed_file(filename):
        # Image files for OCR processing
        IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'tiff', 'bmp', 'gif'}
        # Document files for project documents
        DOCUMENT_EXTENSIONS = {'pdf', 'doc', 'docx', 'txt', 'rtf', 'odt'}
        # Spreadsheet files for data
        SPREADSHEET_EXTENSIONS = {'xls', 'xlsx', 'csv', 'ods'}
        # Presentation files
        PRESENTATION_EXTENSIONS = {'ppt', 'pptx', 'odp'}
        # Bioinformatics and sequence files
        BIOINFORMATICS_EXTENSIONS = {'fastq', 'fq', 'fasta', 'fa', 'fas', 'fna', 'ffn', 'frn', 'sam', 'bam', 'vcf', 'gff', 'gtf', 'bed', 'wig', 'bigwig', 'bw'}
        # Other common project files
        OTHER_EXTENSIONS = {'zip', 'rar', '7z', 'tar', 'gz'}
        
        ALLOWED_EXTENSIONS = IMAGE_EXTENSIONS | DOCUMENT_EXTENSIONS | SPREADSHEET_EXTENSIONS | PRESENTATION_EXTENSIONS | BIOINFORMATICS_EXTENSIONS | OTHER_EXTENSIONS
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
    def get_file_type(filename):
        """Determine file type category"""
        if not filename or '.' not in filename:
            return 'unknown'
            
        ext = filename.rsplit('.', 1)[1].lower()
        
        if ext in {'png', 'jpg', 'jpeg', 'tiff', 'bmp', 'gif'}:
            return 'image'
        elif ext in {'pdf', 'doc', 'docx', 'txt', 'rtf', 'odt'}:
            return 'document'
        elif ext in {'xls', 'xlsx', 'csv', 'ods'}:
            return 'spreadsheet'
        elif ext in {'ppt', 'pptx', 'odp'}:
            return 'presentation'
        elif ext in {'fastq', 'fq', 'fasta', 'fa', 'fas', 'fna', 'ffn', 'frn', 'sam', 'bam', 'vcf', 'gff', 'gtf', 'bed', 'wig', 'bigwig', 'bw'}:
            return 'bioinformatics'
        elif ext in {'zip', 'rar', '7z', 'tar', 'gz'}:
            return 'archive'
        else:
            return 'other'
    
    # Helper function for amplicon detection
    def detect_amplicon_in_file(fastq_path, saved_amplicons, target_amplicon):
        """
        Analyze FASTQ file to detect which amplicon it likely contains
        Returns best matching amplicon info
        """
        import gzip
        from difflib import SequenceMatcher
        # Simple FASTQ parser instead of BioPython dependency
        
        try:
            # Sample first 1000 reads to check for amplicon matches
            sample_sequences = []
            
            # Handle both compressed and uncompressed FASTQ files
            if fastq_path.endswith('.gz'):
                file_handle = gzip.open(fastq_path, 'rt')
            else:
                file_handle = open(fastq_path, 'r')
            
            try:
                seq_count = 0
                line_count = 0
                for line in file_handle:
                    line_count += 1
                    # FASTQ format: every 4 lines is one read (header, sequence, +, quality)
                    if line_count % 4 == 2:  # Sequence line
                        sample_sequences.append(line.strip())
                        seq_count += 1
                        if seq_count >= 1000:  # Sample first 1000 sequences
                            break
            finally:
                file_handle.close()
            
            if not sample_sequences:
                return {'match': 'unknown', 'confidence': 0.0, 'reason': 'No sequences found'}
            
            # Check against target amplicon first
            target_matches = 0
            for seq in sample_sequences[:100]:  # Check first 100 sequences
                similarity = SequenceMatcher(None, seq, target_amplicon).ratio()
                if similarity > 0.7:  # 70% similarity threshold
                    target_matches += 1
            
            target_confidence = target_matches / min(len(sample_sequences), 100)
            
            # Check against saved amplicons
            best_match = {
                'amplicon_name': 'target',
                'confidence': target_confidence,
                'matches': target_matches
            }
            
            for saved_amp in saved_amplicons:
                matches = 0
                for seq in sample_sequences[:100]:
                    similarity = SequenceMatcher(None, seq, saved_amp['amplicon_seq']).ratio()
                    if similarity > 0.7:
                        matches += 1
                
                confidence = matches / min(len(sample_sequences), 100)
                if confidence > best_match['confidence']:
                    best_match = {
                        'amplicon_name': saved_amp['output_name'],
                        'confidence': confidence,
                        'matches': matches
                    }
            
            return {
                'match': best_match['amplicon_name'],
                'confidence': best_match['confidence'],
                'matches': best_match['matches'],
                'total_checked': min(len(sample_sequences), 100)
            }
            
        except Exception as e:
            return {'match': 'error', 'confidence': 0.0, 'reason': str(e)}

    # Test endpoint for debugging JWT
    @app.route('/api/toolbox/test', methods=['GET'])
    @jwt_required()
    def api_toolbox_test():
        """Test endpoint to verify JWT authentication"""
        user_id = int(get_jwt_identity())
        return jsonify({'success': True, 'user_id': user_id, 'message': 'JWT authentication working'})
    
    # Sequence Management API endpoints
    @app.route('/api/projects/<int:project_id>/sequences', methods=['GET'])
    @jwt_required()
    def get_project_sequences(project_id):
        """Get all sequences for a project"""
        user_id = int(get_jwt_identity())
        
        try:
            # Check if user has access to the project
            project = Project.query.filter_by(id=project_id).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
            
            # Check if user is owner or member
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=project_id, user_id=user_id).first()
                if not member:
                    return jsonify({'error': 'Access denied'}), 403
            
            sequences = Sequence.query.filter_by(project_id=project_id).order_by(Sequence.created_at.desc()).all()
            return jsonify({
                'success': True,
                'sequences': [seq.to_dict() for seq in sequences]
            })
            
        except Exception as e:
            return jsonify({'error': f'Failed to fetch sequences: {str(e)}'}), 500
    
    @app.route('/api/sequences', methods=['POST'])
    @jwt_required()
    def create_sequence():
        """Create a new sequence"""
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        try:
            # Validate required fields
            required_fields = ['name', 'sequence_type', 'project_id']
            for field in required_fields:
                if not data.get(field):
                    return jsonify({'error': f'Missing required field: {field}'}), 400
            
            # Check project access
            project = Project.query.filter_by(id=data['project_id']).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
            
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=data['project_id'], user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin', 'member']:
                    return jsonify({'error': 'Access denied'}), 403
            
            # Create sequence
            sequence = Sequence(
                name=data['name'],
                sequence_type=data['sequence_type'],
                sequence_data=data.get('sequence_data', ''),
                description=data.get('description'),
                project_id=data['project_id'],
                user_id=user_id,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc)
            )
            
            # Analyze sequence if data provided
            if sequence.sequence_data:
                sequence.analyze_sequence()
            
            db.session.add(sequence)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'sequence': sequence.to_dict(),
                'message': 'Sequence created successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to create sequence: {str(e)}'}), 500
    
    @app.route('/api/sequences/upload', methods=['POST'])
    @jwt_required()
    def upload_sequence_file():
        """Upload FASTQ/FASTA file and create sequence"""
        user_id = int(get_jwt_identity())
        
        try:
            if 'file' not in request.files:
                return jsonify({'error': 'No file provided'}), 400
            
            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            project_id = request.form.get('project_id')
            sequence_type = request.form.get('sequence_type', 'FASTQ')
            name = request.form.get('name', file.filename.split('.')[0])
            
            if not project_id:
                return jsonify({'error': 'Project ID required'}), 400
            
            # Check project access
            project = Project.query.filter_by(id=int(project_id)).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
            
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=int(project_id), user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin', 'member']:
                    return jsonify({'error': 'Access denied'}), 403
            
            # Save file
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"{timestamp}_{filename}"
            
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'sequences')
            os.makedirs(upload_dir, exist_ok=True)
            file_path = os.path.join(upload_dir, unique_filename)
            
            file.save(file_path)
            file_size = os.path.getsize(file_path)
            
            # Read file content
            sequence_data = ''
            file_format = filename.split('.')[-1].lower()
            
            try:
                if file_format in ['gz', 'gzip']:
                    import gzip
                    with gzip.open(file_path, 'rt') as f:
                        sequence_data = f.read()
                    file_format = 'fastq.gz'
                else:
                    with open(file_path, 'r') as f:
                        sequence_data = f.read()
            except Exception as e:
                return jsonify({'error': f'Failed to read file: {str(e)}'}), 400
            
            # Create sequence record
            sequence = Sequence(
                name=name,
                sequence_type=sequence_type,
                sequence_data=sequence_data,
                original_filename=filename,
                file_format=file_format,
                file_size=file_size,
                project_id=int(project_id),
                user_id=user_id,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc)
            )
            
            # Analyze sequence
            sequence.analyze_sequence()
            
            db.session.add(sequence)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'sequence': sequence.to_dict(),
                'message': 'Sequence file uploaded and analyzed successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Upload failed: {str(e)}'}), 500
    
    @app.route('/api/sequences/<int:sequence_id>', methods=['PUT'])
    @jwt_required()
    def update_sequence(sequence_id):
        """Update an existing sequence"""
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        try:
            sequence = Sequence.query.filter_by(id=sequence_id).first()
            if not sequence:
                return jsonify({'error': 'Sequence not found'}), 404
            
            # Check access
            project = Project.query.filter_by(id=sequence.project_id).first()
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=sequence.project_id, user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin', 'member']:
                    return jsonify({'error': 'Access denied'}), 403
            
            # Update fields
            if 'name' in data:
                sequence.name = data['name']
            if 'description' in data:
                sequence.description = data['description']
            if 'sequence_data' in data:
                sequence.sequence_data = data['sequence_data']
                sequence.analyze_sequence()  # Re-analyze if sequence data changed
            
            sequence.updated_at = datetime.now(timezone.utc)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'sequence': sequence.to_dict(),
                'message': 'Sequence updated successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to update sequence: {str(e)}'}), 500
    
    @app.route('/api/sequences/<int:sequence_id>', methods=['DELETE'])
    @jwt_required()
    def delete_sequence(sequence_id):
        """Delete a sequence"""
        user_id = int(get_jwt_identity())
        
        try:
            sequence = Sequence.query.filter_by(id=sequence_id).first()
            if not sequence:
                return jsonify({'error': 'Sequence not found'}), 404
            
            # Check access
            project = Project.query.filter_by(id=sequence.project_id).first()
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=sequence.project_id, user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin']:
                    return jsonify({'error': 'Access denied'}), 403
            
            db.session.delete(sequence)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Sequence deleted successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to delete sequence: {str(e)}'}), 500
    
    # Amplicon Management API endpoints
    @app.route('/api/projects/<int:project_id>/amplicons', methods=['GET'])
    @jwt_required()
    def get_project_amplicons(project_id):
        """Get all amplicons for a project"""
        user_id = int(get_jwt_identity())
        
        try:
            # Check project access
            project = Project.query.filter_by(id=project_id).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
            
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=project_id, user_id=user_id).first()
                if not member:
                    return jsonify({'error': 'Access denied'}), 403
            
            amplicons = Amplicon.query.filter_by(project_id=project_id).order_by(Amplicon.created_at.desc()).all()
            return jsonify({
                'success': True,
                'amplicons': [amp.to_dict() for amp in amplicons]
            })
            
        except Exception as e:
            return jsonify({'error': f'Failed to fetch amplicons: {str(e)}'}), 500
    
    @app.route('/api/amplicons', methods=['POST'])
    @jwt_required()
    def create_amplicon():
        """Create a new amplicon"""
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        try:
            # Validate required fields
            required_fields = ['name', 'project_id']
            for field in required_fields:
                if not data.get(field):
                    return jsonify({'error': f'Missing required field: {field}'}), 400
            
            # Check project access
            project = Project.query.filter_by(id=data['project_id']).first()
            if not project:
                return jsonify({'error': 'Project not found'}), 404
            
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=data['project_id'], user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin', 'member']:
                    return jsonify({'error': 'Access denied'}), 403
            
            # Create amplicon
            amplicon = Amplicon(
                name=data['name'],
                primer_forward=data.get('primer_forward'),
                primer_reverse=data.get('primer_reverse'),
                target_region=data.get('target_region'),
                expected_size=data.get('expected_size'),
                amplicon_sequence=data.get('amplicon_sequence'),
                annealing_temp=data.get('annealing_temp'),
                cycle_count=data.get('cycle_count'),
                sequence_id=data.get('sequence_id'),
                project_id=data['project_id'],
                user_id=user_id,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc)
            )
            
            # Analyze amplicon if sequence data provided
            if amplicon.amplicon_sequence and amplicon.primer_forward and amplicon.primer_reverse:
                amplicon.analyze_amplicon()
            
            db.session.add(amplicon)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'amplicon': amplicon.to_dict(),
                'message': 'Amplicon created successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to create amplicon: {str(e)}'}), 500
    
    @app.route('/api/amplicons/<int:amplicon_id>', methods=['PUT'])
    @jwt_required()
    def update_amplicon(amplicon_id):
        """Update an existing amplicon"""
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        try:
            amplicon = Amplicon.query.filter_by(id=amplicon_id).first()
            if not amplicon:
                return jsonify({'error': 'Amplicon not found'}), 404
            
            # Check access
            project = Project.query.filter_by(id=amplicon.project_id).first()
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=amplicon.project_id, user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin', 'member']:
                    return jsonify({'error': 'Access denied'}), 403
            
            # Update fields
            updatable_fields = [
                'name', 'primer_forward', 'primer_reverse', 'target_region', 
                'expected_size', 'amplicon_sequence', 'annealing_temp', 'cycle_count'
            ]
            
            for field in updatable_fields:
                if field in data:
                    setattr(amplicon, field, data[field])
            
            # Re-analyze if relevant fields changed
            if any(field in data for field in ['amplicon_sequence', 'primer_forward', 'primer_reverse']):
                amplicon.analyze_amplicon()
            
            amplicon.updated_at = datetime.now(timezone.utc)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'amplicon': amplicon.to_dict(),
                'message': 'Amplicon updated successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to update amplicon: {str(e)}'}), 500
    
    @app.route('/api/amplicons/<int:amplicon_id>', methods=['DELETE'])
    @jwt_required()
    def delete_amplicon(amplicon_id):
        """Delete an amplicon"""
        user_id = int(get_jwt_identity())
        
        try:
            amplicon = Amplicon.query.filter_by(id=amplicon_id).first()
            if not amplicon:
                return jsonify({'error': 'Amplicon not found'}), 404
            
            # Check access
            project = Project.query.filter_by(id=amplicon.project_id).first()
            if project.owner_id != user_id:
                member = ProjectMember.query.filter_by(project_id=amplicon.project_id, user_id=user_id).first()
                if not member or member.role not in ['owner', 'admin']:
                    return jsonify({'error': 'Access denied'}), 403
            
            db.session.delete(amplicon)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': 'Amplicon deleted successfully'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Failed to delete amplicon: {str(e)}'}), 500

    # Toolbox API endpoints for data analysis
    @app.route('/api/toolbox/sequences', methods=['GET', 'POST'])
    @jwt_required()
    def api_toolbox_sequences():
        """Manage amplicon and sgRNA sequences for analysis"""
        import sqlite3
        import os
        
        current_app.logger.info(f"Toolbox sequences endpoint called - Method: {request.method}")
        current_app.logger.info(f"Headers: {dict(request.headers)}")
        
        # Get user ID from JWT
        user_id = int(get_jwt_identity())
        
        # Database path for sequences
        db_path = os.path.join(os.path.dirname(__file__), 'sequences.db')
        current_app.logger.info(f"Database path: {db_path}")
        
        if request.method == 'GET':
            try:
                # Ensure database and table exist
                if not os.path.exists(db_path):
                    return jsonify({'sequences': []})
                
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                
                # Create table if it doesn't exist
                c.execute('''CREATE TABLE IF NOT EXISTS sequences
                           (id INTEGER PRIMARY KEY AUTOINCREMENT,
                            output_name TEXT UNIQUE NOT NULL, 
                            amplicon_seq TEXT NOT NULL, 
                            sgRNA TEXT NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
                
                # Try with created_at first, fallback to without ordering if column doesn't exist
                try:
                    c.execute("SELECT output_name, amplicon_seq, sgRNA FROM sequences ORDER BY created_at DESC")
                except sqlite3.OperationalError:
                    c.execute("SELECT output_name, amplicon_seq, sgRNA FROM sequences")
                sequences = [{'output_name': row[0], 'amplicon_seq': row[1], 'sgRNA': row[2]} 
                           for row in c.fetchall()]
                conn.close()
                return jsonify({'sequences': sequences})
            except Exception as e:
                current_app.logger.error(f"Failed to fetch sequences: {str(e)}")
                return jsonify({'error': f'Failed to fetch sequences: {str(e)}'}), 500
        
        elif request.method == 'POST':
            data = request.get_json()
            output_name = data.get('output_name')
            amplicon_seq = data.get('amplicon_seq')
            sgRNA = data.get('sgRNA')
            
            if not all([output_name, amplicon_seq, sgRNA]):
                return jsonify({'error': 'All fields (output_name, amplicon_seq, sgRNA) are required'}), 400
            
            try:
                # Ensure directory exists for database
                db_dir = os.path.dirname(db_path)
                os.makedirs(db_dir, exist_ok=True)
                
                # Create table if it doesn't exist
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute('''CREATE TABLE IF NOT EXISTS sequences
                           (id INTEGER PRIMARY KEY AUTOINCREMENT,
                            output_name TEXT UNIQUE NOT NULL, 
                            amplicon_seq TEXT NOT NULL, 
                            sgRNA TEXT NOT NULL,
                            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
                
                # Insert or replace sequence
                c.execute("INSERT OR REPLACE INTO sequences (output_name, amplicon_seq, sgRNA) VALUES (?, ?, ?)",
                         (output_name, amplicon_seq, sgRNA))
                conn.commit()
                conn.close()
                
                current_app.logger.info(f"Sequence saved: {output_name}")
                return jsonify({'success': True, 'message': 'Sequence saved successfully'})
            except Exception as e:
                current_app.logger.error(f"Failed to save sequence: {str(e)}")
                return jsonify({'error': f'Failed to save sequence: {str(e)}'}), 500
    
    @app.route('/api/toolbox/analysis', methods=['POST'])
    @jwt_required()
    def api_toolbox_analysis():
        """Start MiSeq data analysis with auto amplicon detection"""
        import subprocess
        import tempfile
        import pandas as pd
        from pathlib import Path
        import uuid
        import sqlite3
        import re
        import gzip
        
        user_id = int(get_jwt_identity())
        data = request.get_json()
        
        # Required parameters
        sample_data = data.get('sample_data', [])  # List of {sample_name, file_name}
        output_name = data.get('output_name')
        amplicon_seq = data.get('amplicon_seq')
        sgRNA = data.get('sgRNA')
        
        # Optional parameters with defaults
        wc = data.get('wc', -10)
        w = data.get('w', 20)
        p = data.get('p', 4)
        base_editor_output = data.get('base_editor_output', True)
        
        if not all([sample_data, output_name, amplicon_seq, sgRNA]):
            return jsonify({'error': 'Missing required parameters'}), 400
        
        try:
            # Create a unique job ID
            job_id = str(uuid.uuid4())
            
            # Get all saved sequences for amplicon detection
            db_path = os.path.join(os.path.dirname(__file__), 'sequences.db')
            saved_amplicons = []
            try:
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute("SELECT output_name, amplicon_seq, sgRNA FROM sequences")
                saved_amplicons = [{'output_name': row[0], 'amplicon_seq': row[1], 'sgRNA': row[2]} 
                                 for row in c.fetchall()]
                conn.close()
            except:
                pass  # Continue without saved sequences if DB doesn't exist
            
            # Create user-specific analysis directory with amplicon-based organization
            analysis_base_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'analysis', str(user_id))
            amplicon_name = output_name.replace(' ', '_').replace('/', '_')
            analysis_dir = os.path.join(analysis_base_dir, amplicon_name)
            os.makedirs(analysis_dir, exist_ok=True)
            
            # Create subdirectories for different analysis outputs
            raw_data_dir = os.path.join(analysis_dir, '01_raw_data')
            processed_dir = os.path.join(analysis_dir, '02_processed')
            results_dir = os.path.join(analysis_dir, '03_results')
            plots_dir = os.path.join(analysis_dir, '04_plots')
            
            for dir_path in [raw_data_dir, processed_dir, results_dir, plots_dir]:
                os.makedirs(dir_path, exist_ok=True)
            
            # Auto-detect amplicon matches in uploaded files
            upload_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'toolbox', str(user_id))
            amplicon_matches = {}
            
            for sample in sample_data:
                file_path = os.path.join(upload_dir, sample['file_name'])
                if os.path.exists(file_path):
                    # Analyze FASTQ file to detect potential amplicon matches
                    detected_amplicon = detect_amplicon_in_file(file_path, saved_amplicons, amplicon_seq)
                    amplicon_matches[sample['sample_name']] = detected_amplicon
                    
                    # Copy file to appropriate amplicon folder
                    dest_file = os.path.join(raw_data_dir, sample['file_name'])
                    import shutil
                    shutil.copy2(file_path, dest_file)
            
            # Create job record with enhanced metadata
            from models import ProcessingJob
            job = ProcessingJob(
                id=job_id,
                job_type='miseq_analysis',
                user_id=user_id,
                status='processing',
                created_at=datetime.now(timezone.utc),
                metadata={
                    'output_name': output_name,
                    'amplicon_seq': amplicon_seq,
                    'sgRNA': sgRNA,
                    'analysis_dir': analysis_dir,
                    'amplicon_matches': amplicon_matches,
                    'parameters': {
                        'wc': wc, 'w': w, 'p': p, 
                        'base_editor_output': base_editor_output
                    }
                }
            )
            db.session.add(job)
            db.session.commit()
            
            # Queue the analysis job for background processing
            # This would typically use Celery in production
            
            return jsonify({
                'success': True,
                'job_id': job_id,
                'analysis_dir': analysis_dir,
                'amplicon_matches': amplicon_matches,
                'message': f'Analysis job started for amplicon: {amplicon_name}'
            })
            
        except Exception as e:
            current_app.logger.error(f"Analysis error: {str(e)}")
            return jsonify({'error': f'Failed to start analysis: {str(e)}'}), 500
    
    @app.route('/api/toolbox/jobs/<job_id>', methods=['GET'])
    @jwt_required()
    def api_toolbox_job_status(job_id):
        """Get status of analysis job"""
        user_id = int(get_jwt_identity())
        
        try:
            from models import ProcessingJob
            job = ProcessingJob.query.filter_by(id=job_id, user_id=user_id).first()
            
            if not job:
                return jsonify({'error': 'Job not found'}), 404
            
            return jsonify({
                'job': job.to_dict()
            })
            
        except Exception as e:
            return jsonify({'error': f'Failed to get job status: {str(e)}'}), 500
    
    @app.route('/api/toolbox/upload-data', methods=['POST'])
    @jwt_required()
    def api_toolbox_upload_data():
        """Upload FASTQ files and Excel metadata for analysis"""
        user_id = int(get_jwt_identity())
        
        # Check if files are present
        if 'files' not in request.files:
            return jsonify({'error': 'No files provided'}), 400
        
        files = request.files.getlist('files')
        if not files or all(f.filename == '' for f in files):
            return jsonify({'error': 'No files selected'}), 400
        
        try:
            # Create user-specific upload directory
            upload_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'toolbox', str(user_id))
            os.makedirs(upload_dir, exist_ok=True)
            
            uploaded_files = []
            
            for file in files:
                if file.filename:
                    # Secure the filename
                    from werkzeug.utils import secure_filename
                    filename = secure_filename(file.filename)
                    
                    # Add timestamp to avoid conflicts
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    name, ext = os.path.splitext(filename)
                    unique_filename = f"{timestamp}_{name}{ext}"
                    
                    file_path = os.path.join(upload_dir, unique_filename)
                    file.save(file_path)
                    
                    uploaded_files.append({
                        'original_name': filename,
                        'stored_name': unique_filename,
                        'file_path': file_path,
                        'file_size': os.path.getsize(file_path)
                    })
            
            return jsonify({
                'success': True,
                'uploaded_files': uploaded_files,
                'message': f'Successfully uploaded {len(uploaded_files)} files'
            })
            
        except Exception as e:
            return jsonify({'error': f'Upload failed: {str(e)}'}), 500

    @app.route('/api/toolbox/miseq', methods=['POST'])
    @jwt_required()
    def api_toolbox_miseq_launcher():
        """Launch MiSeq FastQ Analyzer GUI"""
        import subprocess
        import threading
        
        try:
            user_id = int(get_jwt_identity())
            current_app.logger.info(f"User {user_id} requesting MiSeq analyzer launch")
            
            # Launch the MiSeq GUI in a separate thread to avoid blocking
            def launch_miseq_gui():
                try:
                    miseq_script_path = os.path.join(os.path.dirname(__file__), 'miseq.py')
                    if not os.path.exists(miseq_script_path):
                        current_app.logger.error(f"MiSeq script not found at {miseq_script_path}")
                        return
                    
                    # Launch the GUI using Python
                    subprocess.Popen([
                        'python3', miseq_script_path
                    ], cwd=os.path.dirname(__file__))
                    
                    current_app.logger.info("MiSeq analyzer GUI launched successfully")
                
                except Exception as e:
                    current_app.logger.error(f"Failed to launch MiSeq GUI: {str(e)}")
            
            # Start the launcher in background
            threading.Thread(target=launch_miseq_gui, daemon=True).start()
            
            return jsonify({
                'success': True,
                'message': 'MiSeq FastQ Analyzer is launching. Please check your desktop for the application window.'
            })
            
        except Exception as e:
            current_app.logger.error(f"MiSeq launcher error: {str(e)}")
            return jsonify({'error': f'Failed to launch MiSeq analyzer: {str(e)}'}), 500

    # Error handlers
    # NGS Analysis API endpoints
    @app.route('/api/ngs/analyze', methods=['POST'])
    @jwt_required()
    def api_ngs_analyze():
        """Run NGS analysis on uploaded FASTQ files"""
        try:
            data = request.get_json()
            
            # Import web-safe workflow (no bash commands)
            from ngs_web_workflow import run_web_workflow
            
            # Validate required fields
            required_fields = ['gene_name', 'amplicon_seq', 'sgRNA', 'sample_names']
            for field in required_fields:
                if field not in data:
                    return jsonify({'error': f'Missing required field: {field}'}), 400
            
            # Check if batch_content is provided
            batch_content = data.get('batch_content', None)
            
            # Run the web-safe workflow with optional batch content
            result = run_web_workflow(
                gene_name=data['gene_name'],
                amplicon_seq=data['amplicon_seq'],
                sgRNA=data['sgRNA'],
                sample_names=data['sample_names'],
                batch_content=batch_content
            )
            
            if result['success']:
                return jsonify({
                    'success': True,
                    'message': result['message'],
                    'output_folder': result['gene_folder'],
                    'note': result['note'],
                    'files': {
                        'batch_file': result['batch_file'],
                        'config_file': result['config_file'],
                        'analysis_script': result['analysis_script'],
                        'indel_template': result['indel_template'],
                        'summary_file': result['summary_file'],
                        'instructions_html': result['instructions_html']
                    }
                })
            else:
                return jsonify({'error': 'NGS analysis failed'}), 500
                
        except Exception as e:
            current_app.logger.error(f"NGS analysis error: {str(e)}")
            return jsonify({'error': f'NGS analysis failed: {str(e)}'}), 500
    
    @app.route('/api/ngs/instructions/<gene_name>', methods=['GET'])
    @jwt_required()
    def api_ngs_instructions(gene_name):
        """Get HTML instructions for manual CRISPResso execution"""
        try:
            from pathlib import Path
            
            desktop = Path.home() / "Desktop"
            gene_folder = desktop / gene_name
            instructions_file = gene_folder / "instructions.html"
            
            if not instructions_file.exists():
                return jsonify({'error': 'Instructions not found'}), 404
            
            with open(instructions_file, 'r') as f:
                instructions_html = f.read()
            
            return jsonify({
                'success': True,
                'gene_name': gene_name,
                'instructions_html': instructions_html,
                'instructions_file': str(instructions_file)
            })
            
        except Exception as e:
            current_app.logger.error(f"Instructions error: {str(e)}")
            return jsonify({'error': f'Failed to get instructions: {str(e)}'}), 500
    
    @app.route('/api/ngs/status/<gene_name>', methods=['GET'])
    @jwt_required()
    def api_ngs_status(gene_name):
        """Check NGS analysis status"""
        try:
            from pathlib import Path
            
            desktop = Path.home() / "Desktop"
            gene_folder = desktop / gene_name
            
            if not gene_folder.exists():
                return jsonify({
                    'success': False,
                    'status': 'not_found',
                    'message': f'Analysis folder {gene_name} not found'
                }), 404
            
            # Check for various stages
            batch_file = gene_folder / 'crispresso_batch.txt'
            crispresso_folder = gene_folder / 'CRISPRessoBatch_on_crispresso_batch'
            results_folder = gene_folder / 'results'
            
            status = {
                'folder_exists': True,
                'batch_file_ready': batch_file.exists(),
                'crispresso_complete': crispresso_folder.exists(),
                'results_ready': results_folder.exists()
            }
            
            # Determine overall status
            if crispresso_folder.exists():
                overall_status = 'complete'
            elif batch_file.exists():
                overall_status = 'ready_for_crispresso'
            else:
                overall_status = 'in_progress'
            
            return jsonify({
                'success': True,
                'gene_name': gene_name,
                'status': overall_status,
                'details': status,
                'folder_path': str(gene_folder)
            })
            
        except Exception as e:
            current_app.logger.error(f"Status check error: {str(e)}")
            return jsonify({'error': f'Failed to check status: {str(e)}'}), 500
    
    @app.errorhandler(404)
    def not_found(error):
        return render_template('errors/404.html'), 404
    
    @app.errorhandler(500)
    def internal_error(error):
        db.session.rollback()
        return render_template('errors/500.html'), 500
    
    return app

# Create app instance
app = create_app()

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5003)